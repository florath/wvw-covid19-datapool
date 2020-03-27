'''
Import module handling the data from the
  European Centre for Disease Prevention and Control
Web page
'''

import shutil
import os
import requests
import datetime
import hashlib
from google.cloud import firestore

from data_import.lib.utils import import_data_collection, \
    get_available_data_ids, update_metadata

from openpyxl import load_workbook


LOCAL_DATA_DIR = "/tmp/ecdc_xlsx_data"
LOCAL_DATA_FILE = LOCAL_DATA_DIR + "/d.xlsx"
ECDC_URL = "https://www.ecdc.europa.eu/sites/default/files/documents/" \
    "COVID-19-geographic-disbtribution-worldwide-%04d-%02d-%02d.xlsx"


def download_xlsx():
    '''Donwload the xlsx file and write it to the file system.

    Writing to the file system is needed, because the openpyxl does
    not hanlde 'file like object' correctly.
    '''
    shutil.rmtree(LOCAL_DATA_DIR, ignore_errors=True)
    os.mkdir(LOCAL_DATA_DIR)

    now = datetime.datetime.now()

    # Try for the last five days and use the newest.
    for dl_ts in range(0, 5):
        download_url = ECDC_URL % (now.year, now.month, now.day)
        print("Download URL [%s]" % download_url)

        xlsx_file = requests.get(download_url)
        if xlsx_file.ok:
            break

        now -= datetime.timedelta(days=1)

    if not xlsx_file.ok:
        print("No valid file found")
        return False

    with open(LOCAL_DATA_FILE, "wb") as fd:
        fd.write(xlsx_file.content)

    return True


def handle_one_data_line(line):
    '''Converts one data line into json

    Format of the input data:
    0        1    2      3     4      5       6                          7
    DateRep, Day, Month, Year, Cases, Deaths, Countries and territories, GeoId

    DateRep is ignored as it is the same data in a worse format as 1-3.
    '''
    ts = datetime.datetime(year=int(line[3].value), month=int(line[2].value),
                           day=int(line[1].value)).timestamp()

    nd = {
        'timestamp': ts,
        'infected': int(line[4].value),
        'deaths': int(line[5].value),
        'source': 'ecdc_xlsx',
        'original': {
            'location': line[6].value
        },
        'location': {
            'iso-3166-1-alpha2': line[7].value
        }
    }

    sha_str = str(ts) + "".join(map(lambda x: str(x.value), line[1:]))
    return nd, hashlib.sha256(sha_str.encode("utf-8")).hexdigest()


def update_data(environment):
    '''Reads in the data from the xlsx file, converts it and pushes it
    into the DB'''
    wb = load_workbook(LOCAL_DATA_FILE, read_only=True)
    sheet_ranges = wb['COVID-19-geographic-disbtributi']

    db = firestore.Client()
    tab_ref = db.collection(
        "covid19datapool/%s/ecdc_xlsx/data/collection" % environment)

    data_available_ids = get_available_data_ids(tab_ref)
    # Skip header
    iter_rows = sheet_ranges.iter_rows()
    next(iter_rows)
    import_data_collection(iter_rows,
                           handle_one_data_line, tab_ref,
                           data_available_ids)
    update_metadata(db, "data_import/ecdc_xlsx/metadata.json",
                    environment, "ecdc_xlsx")


def import_data_ecdc_xlsx(environment, ignore_errors):
    print("import_data_ecdc_xlsx called [%s] [%s]" %
          (environment, ignore_errors))
    if download_xlsx():
        update_data(environment)
    print("Finished import_data_ecdc_xlsx")


if __name__ == '__main__':
    '''For (local) testing: only update the data'''
    update_data("test")
